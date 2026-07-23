#!/usr/bin/env python3
"""Original XLSX helper for inspection, error scans, and CSV conversion."""

from __future__ import annotations

import argparse
import csv
import json
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


ERROR_TOKENS = ("#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#N/A", "#NUM!", "#NULL!")


def inspect(path: Path) -> None:
    workbook = load_workbook(path, data_only=False, read_only=False)
    sheets = []
    for sheet in workbook.worksheets:
        formula_count = 0
        nonempty = 0
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value is not None:
                    nonempty += 1
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    formula_count += 1
        sheets.append(
            {
                "name": sheet.title,
                "state": sheet.sheet_state,
                "max_row": sheet.max_row,
                "max_column": sheet.max_column,
                "nonempty_cells": nonempty,
                "formulas": formula_count,
                "merged_ranges": len(sheet.merged_cells.ranges),
                "charts": len(sheet._charts),
                "freeze_panes": str(sheet.freeze_panes) if sheet.freeze_panes else None,
            }
        )
    print(json.dumps({"file": str(path.resolve()), "sheets": sheets}, ensure_ascii=False, indent=2))


def formula_scan(path: Path) -> None:
    formula_book = load_workbook(path, data_only=False, read_only=False)
    value_book = load_workbook(path, data_only=True, read_only=False)
    findings = []
    for formula_sheet in formula_book.worksheets:
        value_sheet = value_book[formula_sheet.title]
        for row in formula_sheet.iter_rows():
            for cell in row:
                formula = cell.value
                cached = value_sheet[cell.coordinate].value
                if isinstance(formula, str) and formula.startswith("="):
                    formula_upper = formula.upper()
                    if any(token in formula_upper for token in ERROR_TOKENS):
                        findings.append({"sheet": formula_sheet.title, "cell": cell.coordinate, "kind": "formula_text", "value": formula})
                if isinstance(cached, str) and cached.upper() in ERROR_TOKENS:
                    findings.append({"sheet": formula_sheet.title, "cell": cell.coordinate, "kind": "cached_error", "value": cached})
    print(json.dumps({"file": str(path.resolve()), "errors": findings, "count": len(findings)}, ensure_ascii=False, indent=2))


def csv_to_xlsx(source: Path, output: Path, delimiter: str | None) -> None:
    chosen = delimiter
    if chosen is None:
        chosen = "\t" if source.suffix.lower() == ".tsv" else ","
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Data"
    with source.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.reader(handle, delimiter=chosen)
        for row_index, row in enumerate(reader, start=1):
            for column_index, value in enumerate(row, start=1):
                sheet.cell(row=row_index, column=column_index, value=value)
    if sheet.max_row >= 1:
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F4E78")
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
    for column in range(1, sheet.max_column + 1):
        width = max((len(str(sheet.cell(row=row, column=column).value or "")) for row in range(1, min(sheet.max_row, 200) + 1)), default=8)
        sheet.column_dimensions[get_column_letter(column)].width = min(max(width + 2, 10), 45)
    output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output)
    inspect(output)


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser()
    sub = parser.add_subparsers(dest="command", required=True)
    inspect_cmd = sub.add_parser("inspect")
    inspect_cmd.add_argument("input", type=Path)
    scan_cmd = sub.add_parser("formula-scan")
    scan_cmd.add_argument("input", type=Path)
    convert_cmd = sub.add_parser("csv-to-xlsx")
    convert_cmd.add_argument("input", type=Path)
    convert_cmd.add_argument("--output", required=True, type=Path)
    convert_cmd.add_argument("--delimiter")
    return parser


def main() -> None:
    args = build_parser().parse_args()
    if args.command == "inspect":
        inspect(args.input)
    elif args.command == "formula-scan":
        formula_scan(args.input)
    elif args.command == "csv-to-xlsx":
        csv_to_xlsx(args.input, args.output, args.delimiter)


if __name__ == "__main__":
    main()
